# Excercise
# ------------------------------------------------------------
# Ex: List each company with respective product count
# Ex: List products with inventory less than 10
# Ex: List each company with respective total inventory value
# Exercise Result: { 'AAA Company': 10969059.95, 'BBB Company': 2375499.47, 'ccc Company' :8114363.62 }
# Ex: Write to Spreadsheet: Calculate and write inventory value for each product into spreadsheet



# Different ways to work with files
# Python has several built-in functions for handling files in general io module --> create, read, write
# Python package to work with spreadhsheets specifically more practical functions for spreadsheets specifically easier to use

import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}

# for product_row in range(2, product_list.max_row + 1):
#        supplier_name = product_list.cell(product_row, 4).value
       
       
#        if supplier_name in product_per_supplier:
#            current_prod_number = product_per_supplier[supplier_name]
#            #current_prod_number = product_per_supplier.get(supplier_name)
#            product_per_supplier[supplier_name] = current_prod_number + 1
#        else :
#            product_per_supplier[supplier_name] = 1

# print(product_per_supplier)    
          

for product_row in range(2, product_list.max_row + 1):  
        
    product_number = product_list.cell(product_row, 1)
    inventory = product_list.cell(product_row, 2)
    
    if int(inventory) < 10:
         print(product_number)
         
